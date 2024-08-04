from ortools.linear_solver import pywraplp
import pandas as pd
import numpy as np
import openpyxl


def vrp_model():

    # sets
    staff_num = 19
    zone_num = 4
    shift_num = 3
    day_num = 31

    # range
    STAFF_NUM = range(staff_num) # root node
    ZONE_NUM = range(zone_num) # destination node
    SHIFT_NUM = range(shift_num) # set of vehicle
    DAY_NUM = range(day_num) # set of compartment

    # parameters
    zone_lst = ["Cargo", "Pax", "Vehicles", "Train", "Cargo"]
    zone_dict = {"Cargo": 0, "Pax": 1, "Vehicles": 2, "Train": 3, np.nan:-1}

    # df_schedule = pd.read_excel("Question 3-Template.xlsx", "Q1 Answer")
    df_schedule = pd.read_excel("Question 3-Template.xlsx", 
                                sheet_name="Q1 Answer", 
                                usecols="C:AG",   
                                skiprows=range(0, 1), 
                                nrows = staff_num)
    
    ava = dict()
    shift_dict = {"O": -1, "M": 0, "A": 1, "N": 2}
    for i in STAFF_NUM:
        for d in DAY_NUM:
            shift = df_schedule.iloc[i, d]
            for s in SHIFT_NUM:
                if shift_dict[shift] == s: 
                    ava[(i, d, s)] = 1
                else:
                    ava[(i, d, s)] = 0
    
    # df_demand = pd.read_excel("Question 3-Template.xlsx", "Q1 Answer")
    df_demand = pd.read_excel("Question 3-Template.xlsx", 
                              sheet_name="Description", 
                              usecols="C:AG",   
                              skiprows=range(0, 2), 
                              nrows = staff_num)
    demand = dict()
    for z in ZONE_NUM:
        for s in SHIFT_NUM:
            for d in DAY_NUM:   
                for s in SHIFT_NUM:
                    demand[(z, d, s)] = df_demand.iloc[z*shift_num + s, d]

    solver = pywraplp.Solver.CreateSolver("SCIP")
    if not solver:
        return
    pass

    # variables
    zone_assign = dict()
    for i in STAFF_NUM:
        for z in ZONE_NUM:
                zone_assign[(i, z)] = solver.BoolVar("")

    # constraints
    # ct2
    for d in DAY_NUM:
        for s in SHIFT_NUM:
            for z in ZONE_NUM:
                solver.Add(solver.Sum([ava[(i, d, s)] * zone_assign[(i, z)] for i in STAFF_NUM]) >= demand[(z, d, s)], "ct1")

    # ct3
    for i in STAFF_NUM:
        solver.Add(solver.Sum([zone_assign[(i, z)] for z in ZONE_NUM]) == 1, "ct3")

    
    # Sets a time limit of 1 hour.
    solver.SetTimeLimit(30*60*1000)

    # set a minimum gap limit for the integer solution during branch and cut
    gap = 0.01
    solverParams = pywraplp.MPSolverParameters()
    solverParams.SetDoubleParam(solverParams.RELATIVE_MIP_GAP, gap)

    print(f"Solving with {solver.SolverVersion()}")
    status = solver.Solve()

    # print solution

    if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
        # print(f"Total distance = {solver.Objective().Value()}")
        zone1 = list()
        for i in STAFF_NUM:
            for z in ZONE_NUM:
                if zone_assign[(i, z)].solution_value() == 1:
                    zone1.append(z)
    else:
        print("No optimal solution")

    # Question 1 answer:
    df_schedule_1 = df_schedule.copy()
    for i in STAFF_NUM:
        for d in DAY_NUM:
            if df_schedule_1.iloc[i, d] != "O": 
                df_schedule_1.iloc[i, d] += " " + zone_lst[zone1[i]]
    
    # Question 2 answer:
    df_schedule_2 = pd.read_excel("Question 3-Template.xlsx", 
                                sheet_name="Q2 Answer", 
                                usecols="C:AG",   
                                skiprows=range(0, 1), 
                                nrows = staff_num)
    zone2 = pd.read_excel("Question 3-Template.xlsx", 
                          sheet_name="Q2 Answer", 
                          usecols="B",   
                          skiprows=range(0, 81), 
                          nrows = staff_num)
    zone2 = zone2.values.tolist()
    zone2 = [zone_dict[element] for innerList in zone2 for element in innerList]
    print(zone2)
    for i in STAFF_NUM:
        for d in DAY_NUM:
            if df_schedule_2.iloc[i, d] != "O": 
                df_schedule_2.iloc[i, d] += " " + zone_lst[zone2[i]+1]
    
    return df_schedule_1, df_schedule_2

if __name__ == "__main__":
    # Solution for question 1
    df_schedule_1, df_schedule_2 = vrp_model()


    # Load the existing Excel file
    file_path = 'Question 3-Answer.xlsx'
    sheet_name_1 = 'Q1 Answer'
    sheet_name_2 = 'Q2 Answer'
    wb = openpyxl.load_workbook(file_path)
    ws1 = wb[sheet_name_1]
    ws2 = wb[sheet_name_2]

    # Specify the starting cell
    start_row = 3  # Starting row
    start_col = 3  # Starting column (3 corresponds to column 'C')

    # Convert DataFrame to a list of lists
    data_list = df_schedule_1.values.tolist()

    # Write DataFrame to the specified location
    for i, row in enumerate(data_list):
        for j, value in enumerate(row):
            ws1.cell(row=start_row + i, column=start_col + j, value=value)
            ws2.cell(row=start_row + i, column=start_col + j, value=value)

    # Save the workbook
    wb.save(file_path)

